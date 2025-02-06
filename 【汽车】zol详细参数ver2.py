# coding=UTF-8
from struct import pack_into
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.comments import Comment
import random
from time import sleep
from bs4 import BeautifulSoup
import urllib.request
import os
import re
import sys

REstr = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
runpath = "/root/car_zol/"
# 开始
urls = {
    '00': u'https://detail.zol.com.cn/car_advSearch/subcate2530_1_s10965-s10970-s10971-s10972-s10973-s10974-s10975-s10966-s10976-s10977-s10978-s10979-s10980-s10967-s10981-s10982-s10983-s10984-s10968-s10969_1_1_0_',
    '11': u'https://detail.zol.com.cn/car_advSearch/subcate2530_1_s10965-s10970-s10971-s10972-s10973-s10974-s10975-s10966-s10976-s10977-s10978-s10979-s10980-s10967-s10981-s10982-s10983-s10984-s10968-s10969_1_1_',
    # https://detail.zol.com.cn/car_advSearch/subcate2530_1_s10965-s10970-s10971-s10972-s10973-s10974-s10975-s10966-s10976-s10977-s10978-s10979-s10980-s10967-s10981-s10982-s10983-s10984-s10968-s10969_1_1_0_1.html#showc
}
headers = [{
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'}, {'User-Agent': 'Mozilla/5.0 (MSIE 10.0; Windows NT 6.1; Trident/5.0)'}, {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:33.0) Gecko/20120101 Firefox/33.0'}, {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/7046A194A'}]
titles = ['序号(主图)', '品牌型号(详情页)', '参考价格(图库页)', '上市时间', '车型名称', '厂商',
          '级别', '能源类型', '环保标准', '厂商指导价(元)', '最大功率(kW)', '最大扭矩(N·m)', '发动机', '变速箱', '长*宽*高(mm)', '车身结构', '最高车速(km/h)', '官方0-100km/h加速(s)', 'WLTC综合油耗(L/100km)', '整车质保', "长度(mm)", "宽度(mm)", "高度(mm)", "轴距(mm)", "前轮距(mm)", "后轮距(mm)", "满载最小离地间隙(mm)", "接近角(°)", "离去角(°)", "车身结构", "车门开启方式", "油箱容积(L)", "后备厢容积(L)", "整备质量(kg)", "最大满载质量(kg)", "风阻系数(Cd)", "发动机型号", "排量(mL)", "排量(L)", "进气形式", "发动机布局", "气缸排列形式", "气缸数(个)", "每缸气门数(个)", "配气机构", "最大马力(Ps)", "最大功率(kW)", "最大扭矩(N·m)", "最大功率转速(rpm)", "最大扭矩转速(rpm)", "最大净功率(kW)", "燃料形式", "燃油标号", "供油方式", "缸盖材料", "缸体材料", "环保标准", "挡位个数", "变速箱类型", "简称", "驱动方式", "四驱形式", "中央差速器结构", "前悬架类型", "后悬架类型", "助力类型", "车体结构", "前制动器类型", "后制动器类型", "驻车制动类型", "前轮胎规格", "后轮胎规格", "备胎规格", "主/副驾驶座安全气囊", "前/后排侧气囊", "前/后排头部气囊(气帘)", "胎压监测功能", "缺气保用轮胎", "安全带未系提醒", "ISOFIX儿童座椅接口", "ABS防抱死", "制动力分配(EBD/CBC等)", "刹车辅助(EBA/BAS/BA等)", "牵引力控制(ASR/TCS/TRC等)", "车身稳定控制(ESC/ESP/DSC等)", "被动行人保护", "车道偏离预警系统", "主动刹车/主动安全系统", "疲劳驾驶提示", "前方碰撞预警", "道路救援呼叫", "驾驶模式切换", "发动机启停技术", "自动驻车", "上坡辅助", "陡坡缓降", "空气悬架", "可变转向比", "可变悬架功能", "前/后驻车雷达", "驾驶辅助影像", "巡航系统", "辅助驾驶等级", "倒车车侧预警系统", "卫星导航系统", "导航路况信息显示", "并线辅助", "车道保持辅助系统", "车道居中保持", "道路交通标识识别", "夜视系统", "轮圈材质", "电动后备厢", "发动机电子防盗", "车内中控锁", "钥匙类型", "无钥匙启动系统", "无钥匙进入功能", "运动外观套件", "近光灯光源", "远光灯光源", "LED日间行车灯", "自适应远近光", "自动头灯", "大灯高度可调", "大灯延时关闭", "大灯清洗装置", "前大灯雨雾模式", "车前雾灯", "灯光特色功能", "天窗类型", "前/后电动车窗", "车窗一键升降功能", "车窗防夹手功能", "车内化妆镜", "后雨刷", "感应雨刷功能", "后排侧隐私玻璃", "后风挡遮阳帘", "多层隔音玻璃", "外后视镜功能", "中控彩色屏幕", "中控屏幕尺寸", "中控下屏幕尺寸", "后排控制多媒体", "手机互联/映射", "蓝牙/车载电话", "语音识别控制系统", "车载智能系统", "车联网", "4G/5G网络", "OTA升级", "手机APP远程功能", "方向盘材质", "方向盘位置调节", "换挡形式", "多功能方向盘", "方向盘换挡", "方向盘加热", "方向盘记忆", "行车电脑显示屏幕", "全液晶仪表盘", "液晶仪表尺寸", "HUD抬头数字显示", "ETC装置", "内后视镜功能", "多媒体/充电接口", "USB/Type-C接口数量", "手机无线充电功能", "行李厢12V电源接口", "座椅材质", "主座椅调节方式", "副座椅调节方式", "主/副驾驶座电动调节", "前排座椅功能", "电动座椅记忆功能", "副驾驶位后排可调节按钮", "第二排座椅调节", "第二排座椅电动调节", "第二排座椅功能", "后排座椅放倒形式", "前/后中央扶手", "后排杯架", "运动风格座椅", "扬声器品牌名称", "扬声器数量", "触摸式阅读灯", "车内环境氛围灯", "空调温度控制方式", "后排独立空调", "后座出风口", "温度分区控制", "车载空气净化器", "车内PM2.5过滤装置", "负离子发生器", "车内香氛装置", "CLTC纯电续航里程(km)", "电动机(Ps)", "电能当量燃料消耗量(L/100km)", "整车质保", "车身", "最小转弯半径(m)", "前备厢容积(L)", "电动机", "电机类型", "电动机总功率(kW)", "电动机总功率(Ps)", "电动机总扭矩(N·m)", "前电动机最大功率(kW)", "前电动机最大扭矩(N·m)", "驱动电机数", "电机布局", "电池冷却方式", "电池能量(kWh)", "百公里耗电量(kWh/100km)", "高压快充", "快充功能", "后电动机最大功率(kW)", "后电动机最大扭矩(N·m)", "变速箱", "底盘转向", "车轮制动", "被动安全", "前排中间气囊", "主动安全", "后方碰撞预警", "低速行车警告", "内置行车记录仪", "DOW开门预警", "哨兵模式/千里眼", "防侧翻系统", "驾驶操控", "能量回收系统", "驾驶硬件", "透明底盘/540度影像", "摄像头数量", "超声波雷达数量", "毫米波雷达数量", "激光雷达数量", "芯片总算力", "辅助驾驶芯片", "驾驶功能", "辅助驾驶系统", "地图品牌", "自动变道辅助", "匝道自动驶出(入)", "自动驾驶辅助路段", "信号灯识别", "循迹倒车", "方向盘离手检测", "自动泊车入位", "记忆泊车", "遥控泊车", "起步提醒", "远程召唤", "外观/防盗", "电动后备厢位置记忆", "主动闭合式进气格栅", "无框设计车门", "电池预加热", "电动扰流板", "电动吸合车门", "对外放电", "车外灯光", "转向辅助灯", "天窗/玻璃", "外后视镜", "屏幕/系统", "可见即可说", "后排液晶屏幕尺寸", "面部识别", "声纹识别", "语音免唤醒词", "语音分区域唤醒识别", "语音连续识别", "车机智能芯片", "智能化配置", "Wi-Fi热点", "主动降噪", "模拟声浪", "车载KTV", "方向盘/内后视镜", "AR-HUD增强现实抬头显示", "车内充电", "手机无线充电功率", "座椅配置", "音响/车内灯光", "杜比全景声(Dolby Atmos)", "空调/冰箱", "热泵空调", "车载冰箱", "基本参数", "发动机", "WLTC综合油耗(L/100km)", "膝部气囊", "AR实景导航", "感应后备厢", "远程启动功能", "转向头灯", "可加热喷水嘴", "V2X通讯", "首任车主质保政策", "车门数(个)", "座位数(个)", "后排侧窗遮阳帘", "隐藏电动门把手", "副驾娱乐屏", "手势控制", "最大爬坡度(%)", "最大爬坡角度(°)", "换电", "电池能量密度(Wh/kg)", "单踏板模式", "实测0-100km/h加速(s)", "实测100-0km/h制动(m)", "实测油耗(L/100km)", "压缩比", "发动机特有技术", "电池类型", "电池组质保", "快充时间(小时)", "慢充时间(小时)", "快充电量(%)", "系统综合功率(Ps)", "系统综合功率(kW)", "星空天窗", "后排液晶屏幕", "NEDC纯电续航里程(km)", "WLTC纯电续航里程(km)", "官方0-50km/h加速(s)", "最低荷电状态油耗(L/100km)", "快充功率(kW)", "旋转大屏", "缸径(mm)", "行程(mm)", "电芯品牌", "220V/230V电源", "车侧脚踏板", "车顶行李架", "加热/制冷杯架", "车内生物监测系统", "座椅布局", "WLTC综合续航(km)", "四驱/越野", "拖挂钩", "高精地图", "自动开合车门", "场景灯语", "多指飞屏操控", "第二排独立座椅", "第三排座椅调节", "第三排座椅电动调节", "第三排座椅功能", "后排座椅电动放倒", "系统综合扭矩(N·m)", "中控液晶屏分屏显示", "侧滑门形式", "后排小桌板", "纵向通过角(°)", "最大涉水深度(mm)", "准拖挂车总质量(kg)", "限滑差速器/差速锁", "中央差速器锁止功能", "低速四驱", "坦克转弯", "绞盘", "蠕行模式"]
comment_text = ["序号(主图)：可以查看大图", "品牌型号(详情页)：可以跳转到参数详情",
                "参考价格(图库页)：可以跳转到图库", "上市时间：", "车型名称：", "厂商：", "级别："]


def _re(stre):
    r = stre.replace("\n\t\r", "，").replace("\n\r\t", "，").replace("\t\n\r", "，").replace("\t\r\n", "，").replace("\r\n\t", "，").replace("\r\t\n", "，").replace("\n\n", "，").replace("\n\t", "，").replace("\t\n", "，").replace("\n\r", "，").replace("\r\n", "，").replace("\r\n", "，").replace("\t\n", "，").replace("\n\t", "，").replace("\t\t", "，").replace("\t\r", "，").replace("\r\t", "，").replace("\r\n", "，").replace("\n\t", "，").replace("\r\t", "，").replace("\t\r", "，").replace("\r\r", "，").replace("\r", "，").replace("\t", "，").replace("\n", "，").replace(",,", "，").replace(",", "，").replace(
        "，，", "，").replace("\xa0", "").replace("纠错", "").replace(">", "").replace("<", "").replace("查看官方图", "").replace("查看外观图", "").replace("更多商用笔记本", "").replace("更多家用笔记本", "").replace("更多酷睿i9CPU", "").replace("更多酷睿i7CPU", "").replace("更多酷睿i5CPU", "").replace("运行流畅", "").replace("多任务运行强", "").replace("极速运行热门游戏本", "").replace("更多15.6英寸笔记本", "").replace("更多14英寸笔记本", "").replace("更多16英寸笔记本", "").replace("更多16.1英寸笔记本", "").replace("超高清屏笔记本", "").replace("全高清屏笔记本", "").replace("2GB显存笔记本", "").replace("4GB显存笔记本", "").replace("6GB显存笔记本", "").replace("8GB显存笔记本", "").replace("游戏、便捷", "").replace("笔记本", "").replace("更多", "").replace("查看官方图", "").replace("查看外观图", "").replace("更多", "").replace("手机性能排行", "").replace("查看外观", "").replace("游戏运行卡顿", "").replace("游戏运行良好", "").replace("游戏运行一般", "").replace("游戏运行流畅", "").replace("5.2万张照片2.2万首歌曲", "").replace("2.6万张照片1.1万首歌曲", "").replace("10.5万张照片4.4万首歌曲", "").replace("1.3万张照片5461首歌曲", "").replace("3276张照片1365首歌曲", "").replace("进入官网", " ").replace("  ", " ")
    return r


def _spider(year):
    yearl = "在库全级别的の"
    print("开始爬取[ "+yearl+" ]【汽车】精简参数...")
    bi_ = ['=HYPERLINK(""," ")', '=HYPERLINK(""," ")', '=HYPERLINK(""," ")']
    if (os.path.exists(runpath+yearl+'【汽车】精简参数.xlsx') == False):
        # 实例化
        wb = Workbook()
        # 激活 worksheet
        ws = wb.active
        # 添加单元格注释
        for i in range(len(comment_text)):
            cell = ws.cell(row=1, column=(i+1))
            comment = Comment(text=comment_text[i], author="作者")
            cell.comment = comment
        # ws.append(bi_)
        ws.append(titles)
        ws.freeze_panes = 'E3'
        max_columns = ws.max_column  # 获取最大列
        fills = PatternFill("solid", fgColor="FFC0CB")  # 设置填充颜色为粉红色
        for jj in range(1, max_columns + 1):
            ws.cell(2, jj).fill = fills
        FullRange = "A2:" + get_column_letter(ws.max_column) + str(ws.max_row)
        ws.auto_filter.ref = FullRange
        wb.save(runpath+yearl+'【汽车】精简参数.xlsx')
    allurl = urls[year]+"1.html"
    header_i = random.choice([0, 1, 2, 3])
    header = headers[header_i]
    res_all = urllib.request.urlopen(urllib.request.Request(
        url=allurl, headers=header))
    page_all = res_all.read().decode('GB18030')
    soup = BeautifulSoup(page_all, 'html.parser')
    # 获取总页数
    try:
        pages = soup.select("p.page_order")[0].text.replace(
            " ", "").replace("1/", "").replace("下一页", "").replace("\n", "").replace("\t", "")
    except BaseException as e:
        pages = 0
    else:
        if (os.path.exists(runpath+yearl+'【汽车】精简参数.状态.txt') == False):
            f = open(runpath+yearl+'【汽车】精简参数.状态.txt', 'w')
            f.write('0,0')
            f.close()
            pageing = 0
        else:
            f = open(runpath+yearl+'【汽车】精简参数.状态.txt', "r")
            l_f = f.readlines()  # 读取全部内容
            f.close()
            l_f = l_f[0].split(',')
            pageing = int(l_f[0])
        for i in range(pageing, int(pages)+1):  # 按页循环
            if (i > 0):  # 没什么作用
                j = i  # 记录页数
                s_page = i
                if (i == 0):
                    flg = ""
                    url = urls['00']+flg+str(i)+".html#showc"
                else:
                    url = urls['11']+"_"+str(i)+".html#showc"
                header_i = random.choice([0, 1, 2, 3])
                header = headers[header_i]
                res = urllib.request.urlopen(urllib.request.Request(
                    url=url, headers=header))
                page = res.read().decode('GB18030')
                soup = BeautifulSoup(page, 'html.parser')
                f = open(runpath+yearl+'【汽车】精简参数.状态.txt', "r")
                l_f = f.readlines()  # 读取全部内容
                l_f = l_f[0].split(',')
                f.close()
                detail = int(l_f[1])
                for i in range(detail, 30):
                    if (i == 29):
                        s_detail = 0
                    else:
                        s_detail = i
                    l = (j-1)*30 + i + 1  # 记录总数
                    print(" ")
                    print("第 "+str(j)+" 页/共 "+pages +
                          " 页>>总第 "+str(l)+" 台【汽车】："+url)
                    arr = []  # 建立一个数组列表存储一部手机的信息
                    try:
                        name = soup.select(
                            "ul.result_list>li>dl.pro_detail>dt>a")[i].get_text()
                    except BaseException as e:
                        print("获取名称失败："+e)
                        sys.exit()
                    try:
                        href = soup.select("ul.result_list>li>dl.pro_detail>dd>div>ul>li>a")[
                            i]['href']
                    except BaseException as e:
                        print("获取链接："+e)
                        sys.exit()
                    try:
                        pic1 = soup.select("ul.result_list>li>div.check_pic")[
                            i].a.img.get("src")
                    except BaseException as e:
                        pic1 = " "
                    try:
                        price = soup.select("ul.result_list>li>div.date_price")[
                            i].get_text().replace('\n', "")
                    except BaseException as e:
                        price = " "
                    price = price.split('[')
                    price = price[0]
                    arr.append(name)
                    arr.append(href)
                    arr.append(price)
                    header_i = random.choice([0, 1, 2, 3])
                    header = headers[header_i]
                    relhref = "https://detail.zol.com.cn"+href
                    # print(relhref)
                    res1 = urllib.request.urlopen(urllib.request.Request(
                        url=relhref, headers=header))
                    page1 = res1.read()
                    soup1 = BeautifulSoup(page1, 'lxml')
                    # 开始准备数据/
                    # 获取图库
                    try:
                        piclink = soup1.select(".big-pic-fl>a")[0]['href']
                    except BaseException as e:
                        piclink = " "
                    piclink = _re(piclink)
                    if (piclink == " "):
                        relpiclink = relhref.replace(
                            "param.shtml", "pic.shtml")
                    else:
                        relpiclink = "https://detail.zol.com.cn/"+piclink
                    # 获取主图
                    try:
                        imglink = soup1.select(".big-pic-fl>a>img")[0]['src']
                    except BaseException as e:
                        imglink = " "
                    imglink = _re(imglink)
                    if (imglink == " "):
                        imglink = pic1
                    l_ = ['=HYPERLINK("'+REstr.sub(r'', imglink)+'","'+str(l)+'")', '=HYPERLINK("'+REstr.sub(r'', relhref)+'","'+REstr.sub(
                        r'', name)+'")', '=HYPERLINK("'+REstr.sub(r'', relpiclink)+'","'+REstr.sub(r'', price)+'")']
                    newPmName_ = soup1.select("table>tr>th>span")
                    newPmName_dict = {}
                    for i in range(len(newPmName_)):
                        newPmName_dict[newPmName_[i].text] = newPmVal_ = soup1.select(
                            "#newPmVal_"+str(i))[0].text
                    # print(newPmName_dict)
                    for i in range(3, len(titles)):
                        strr = titles[i]
                        if (strr in newPmName_dict):
                            temp = newPmName_dict.get(strr)
                        else:
                            temp = " "
                        temp = REstr.sub(r'', temp)
                        temp = _re(temp)
                        l_.append(temp)
                    wb1 = load_workbook(runpath+yearl+'【汽车】精简参数.xlsx')
                    # 激活 worksheet
                    ws1 = wb1.get_sheet_by_name("Sheet")
                    ws1 = wb1.active
                    ws1.append(l_)
                    wb1.save(runpath+yearl+'【汽车】精简参数.xlsx')
                    wb1.close()
                    a = open(runpath+yearl+'【汽车】精简参数.状态.txt', 'w')
                    a.write(str(s_page)+','+str(s_detail))
                    a.close()
                    print(l_)
                    # sleep_time = random.randint(0, 0)
                    # sleep(sleep_time)
                # sleep_time = random.randint(1, 2)
                # sleep(sleep_time)
                j = j+1


if __name__ == "__main__":
    _spider('00')
